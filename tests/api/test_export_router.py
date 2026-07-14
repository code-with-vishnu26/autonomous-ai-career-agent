"""Phase 65 (ADR-0083): the `/export/*.xlsx` download endpoints.

Each route builds a real openpyxl workbook from the caller's own rows and
streams it back -- these tests load the returned bytes with openpyxl and
assert on the actual sheet, not just the status code, and prove one user
never sees another's rows.
"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from career_agent.api.app import create_app
from career_agent.api.rate_limit import auth_rate_limiter
from career_agent.core.security import (
    create_access_token,
    create_resume_download_token,
    hash_password,
)
from career_agent.domain.application_session import ApplicationSession
from career_agent.domain.models import (
    BasicsSection,
    MasterProfile,
    Opportunity,
    Provenance,
    TailoredContent,
)
from career_agent.domain.resume_variants import ResumeVariant
from career_agent.domain.submission import SubmissionResult
from career_agent.domain.user import User
from career_agent.storage.sqlite import (
    SqliteApplicationSessionStore,
    SqliteMasterProfileStore,
    SqliteOpportunityRepository,
    SqliteResumeVariantStore,
    SqliteSubmissionResultStore,
    SqliteUserStore,
)

_JWT_SECRET = "unit-test-secret-not-for-real-use"
_OWNER_ID = "owner-user-id"
_OTHER_ID = "other-user-id"
_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@pytest.fixture(autouse=True)
def _isolated_database(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    monkeypatch.setenv("DATABASE_PATH", str(tmp_path / "career_agent.db"))
    monkeypatch.setenv("JWT_SECRET_KEY", _JWT_SECRET)


@pytest.fixture(autouse=True)
def _reset_rate_limiter() -> None:
    auth_rate_limiter._hits.clear()
    yield
    auth_rate_limiter._hits.clear()


def _db_path() -> Path:
    return Path("career_agent.db")


def _create_user(user_id: str, email: str) -> str:
    SqliteUserStore(_db_path()).create(
        User(
            id=user_id,
            email=email,
            hashed_password=hash_password("irrelevant-not-checked-here"),
            role="user",
            created_at=datetime(2026, 1, 1, tzinfo=UTC),
        )
    )
    return create_access_token(
        user_id=user_id,
        role="user",
        secret_key=_JWT_SECRET,
        expires_in_minutes=15,
    )


@pytest.fixture
def client() -> TestClient:
    return TestClient(create_app())


def _application_session(**overrides: object) -> ApplicationSession:
    fields: dict[str, object] = {
        "id": "sess-1",
        "provider": "greenhouse",
        "company": "Acme Corp",
        "job_title": "Backend Engineer",
        "url": "https://boards.greenhouse.io/acme/jobs/1",
        "opportunity_id": "opp-1",
        "status": "READY_FOR_REVIEW",
        "created_at": datetime(2026, 1, 1, tzinfo=UTC),
    }
    fields.update(overrides)
    return ApplicationSession(**fields)


def _submission_result(**overrides: object) -> SubmissionResult:
    fields: dict[str, object] = {
        "id": "sub-1",
        "application_session_id": "sess-1",
        "review_session_id": "review-1",
        "opportunity_id": "opp-1",
        "provider": "greenhouse",
        "company": "Acme Corp",
        "job_title": "Backend Engineer",
        "submitted": True,
        "status": "SUBMITTED",
    }
    fields.update(overrides)
    return SubmissionResult(**fields)


def test_applications_export_requires_authentication(client: TestClient) -> None:
    assert client.get("/export/applications.xlsx").status_code == 401


def test_submissions_export_requires_authentication(client: TestClient) -> None:
    assert client.get("/export/submissions.xlsx").status_code == 401


def test_applications_export_returns_a_real_xlsx(client: TestClient) -> None:
    token = _create_user(_OWNER_ID, "owner@example.com")
    SqliteApplicationSessionStore(_db_path()).save(
        _application_session(company="Acme Corp", job_title="Backend Engineer"),
        user_id=_OWNER_ID,
    )
    response = client.get(
        "/export/applications.xlsx", headers={"Authorization": f"Bearer {token}"}
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == _XLSX_MEDIA_TYPE
    assert "applications.xlsx" in response.headers["content-disposition"]

    sheet = load_workbook(BytesIO(response.content)).active
    header = [cell.value for cell in sheet[1]]
    assert "Company" in header
    assert "Role" in header
    values = [cell.value for cell in sheet[2]]
    assert "Acme Corp" in values
    assert "Backend Engineer" in values


async def test_applications_export_is_enriched_with_posting_details_and_research(
    client: TestClient,
) -> None:
    token = _create_user(_OWNER_ID, "owner@example.com")
    SqliteApplicationSessionStore(_db_path()).save(
        _application_session(
            company="Acme Corp",
            opportunity_id="opp-9",
            url="https://boards.greenhouse.io/acme/jobs/9",
            cover_letter_body="Dear Acme, I am excited...",
        ),
        user_id=_OWNER_ID,
    )
    await SqliteOpportunityRepository(_db_path()).add(
        Opportunity(
            id="opp-9",
            company_id="acme",
            canonical_company="Acme Corp",
            title="Backend Engineer",
            source="ats_api",
            source_url="https://boards.greenhouse.io/acme/jobs/9",
            provenance=Provenance(
                method="structured_api",
                reference="https://boards.greenhouse.io/acme/jobs/9",
                extraction_confidence=1.0,
            ),
            location="Remote - India",
            remote=True,
            description_raw="Python role.",
            discovered_at=datetime(2026, 1, 1, tzinfo=UTC),
        )
    )
    response = client.get(
        "/export/applications.xlsx", headers={"Authorization": f"Bearer {token}"}
    )
    assert response.status_code == 200
    sheet = load_workbook(BytesIO(response.content)).active
    header = [cell.value for cell in sheet[1]]
    for expected in ("Job URL", "Careers Page", "Company Research", "Cover Letter"):
        assert expected in header
    values = [cell.value for cell in sheet[2]]
    assert "https://boards.greenhouse.io/acme/jobs/9" in values  # accurate job link
    assert "Remote - India" in values  # accurate location from the opportunity
    assert "Dear Acme, I am excited..." in values  # cover letter inline
    # No search key configured in tests -> honest "no key" note, never fabricated.
    assert any(
        isinstance(v, str) and "no web-search key" in v.lower() for v in values
    )
    # The job URL cell is a real clickable hyperlink.
    job_url_col = header.index("Job URL") + 1
    assert sheet.cell(row=2, column=job_url_col).hyperlink is not None


def test_submissions_export_returns_a_real_xlsx(client: TestClient) -> None:
    token = _create_user(_OWNER_ID, "owner@example.com")
    SqliteSubmissionResultStore(_db_path()).save(
        _submission_result(company="Acme Corp"), user_id=_OWNER_ID
    )
    response = client.get(
        "/export/submissions.xlsx", headers={"Authorization": f"Bearer {token}"}
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == _XLSX_MEDIA_TYPE

    sheet = load_workbook(BytesIO(response.content)).active
    values = [cell.value for cell in sheet[2]]
    assert "Acme Corp" in values
    assert "yes" in values  # submitted rendered readable


def test_empty_export_is_a_valid_header_only_workbook(client: TestClient) -> None:
    token = _create_user(_OWNER_ID, "owner@example.com")
    response = client.get(
        "/export/applications.xlsx", headers={"Authorization": f"Bearer {token}"}
    )
    assert response.status_code == 200
    sheet = load_workbook(BytesIO(response.content)).active
    assert sheet.max_row == 1  # header only, no data rows
    assert [cell.value for cell in sheet[1]][0] == "Prepared"


def _master_profile(**overrides: object) -> MasterProfile:
    fields: dict[str, object] = {
        "version": "pending",
        "basics": BasicsSection(name="Ada Lovelace", email="ada@example.com"),
    }
    fields.update(overrides)
    return MasterProfile(**fields)


def _resume_variant(**overrides: object) -> ResumeVariant:
    fields: dict[str, object] = {
        "id": "variant-1",
        "category": "backend",
        "profile_version": "v1",
        "content": TailoredContent(summary="Backend engineer, 5 years Python."),
        "created_at": "2026-01-01T00:00:00+00:00",
    }
    fields.update(overrides)
    return ResumeVariant(**fields)


def test_applications_export_includes_linkedin_and_resume_pdf_columns(
    client: TestClient,
) -> None:
    token = _create_user(_OWNER_ID, "owner@example.com")
    SqliteMasterProfileStore(_db_path()).save(_OWNER_ID, _master_profile())
    SqliteResumeVariantStore(_db_path()).save(
        _resume_variant(), user_id=_OWNER_ID
    )
    SqliteApplicationSessionStore(_db_path()).save(
        _application_session(resume_variant_id="variant-1"), user_id=_OWNER_ID
    )
    response = client.get(
        "/export/applications.xlsx", headers={"Authorization": f"Bearer {token}"}
    )
    assert response.status_code == 200
    sheet = load_workbook(BytesIO(response.content)).active
    header = [cell.value for cell in sheet[1]]
    assert "Company LinkedIn" in header
    assert "Résumé (PDF)" in header

    resume_col = header.index("Résumé (PDF)") + 1
    cell = sheet.cell(row=2, column=resume_col)
    assert cell.hyperlink is not None
    assert "/export/resume/variant-1.pdf?token=" in cell.hyperlink.target


def test_applications_export_leaves_resume_column_blank_without_a_variant(
    client: TestClient,
) -> None:
    token = _create_user(_OWNER_ID, "owner@example.com")
    SqliteApplicationSessionStore(_db_path()).save(
        _application_session(), user_id=_OWNER_ID
    )
    response = client.get(
        "/export/applications.xlsx", headers={"Authorization": f"Bearer {token}"}
    )
    sheet = load_workbook(BytesIO(response.content)).active
    header = [cell.value for cell in sheet[1]]
    resume_col = header.index("Résumé (PDF)") + 1
    assert sheet.cell(row=2, column=resume_col).value in (None, "")


def test_resume_pdf_download_returns_a_real_pdf(client: TestClient) -> None:
    _create_user(_OWNER_ID, "owner@example.com")
    SqliteMasterProfileStore(_db_path()).save(_OWNER_ID, _master_profile())
    SqliteResumeVariantStore(_db_path()).save(
        _resume_variant(), user_id=_OWNER_ID
    )
    token = create_resume_download_token(
        user_id=_OWNER_ID,
        resume_variant_id="variant-1",
        secret_key=_JWT_SECRET,
        expires_in_days=90,
    )
    # No Authorization header at all -- the link token is the only auth.
    response = client.get(f"/export/resume/variant-1.pdf?token={token}")
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert response.content.startswith(b"%PDF")


def test_resume_pdf_download_rejects_a_garbage_token(client: TestClient) -> None:
    _create_user(_OWNER_ID, "owner@example.com")
    response = client.get("/export/resume/variant-1.pdf?token=not-a-real-jwt")
    assert response.status_code == 401


def test_resume_pdf_download_rejects_a_token_for_a_different_variant(
    client: TestClient,
) -> None:
    _create_user(_OWNER_ID, "owner@example.com")
    SqliteMasterProfileStore(_db_path()).save(_OWNER_ID, _master_profile())
    SqliteResumeVariantStore(_db_path()).save(
        _resume_variant(id="variant-1"), user_id=_OWNER_ID
    )
    token = create_resume_download_token(
        user_id=_OWNER_ID,
        resume_variant_id="variant-2",  # different from the URL below
        secret_key=_JWT_SECRET,
        expires_in_days=90,
    )
    response = client.get(f"/export/resume/variant-1.pdf?token={token}")
    assert response.status_code == 401


def test_resume_pdf_download_rejects_another_users_token(client: TestClient) -> None:
    _create_user(_OWNER_ID, "owner@example.com")
    _create_user(_OTHER_ID, "other@example.com")
    SqliteMasterProfileStore(_db_path()).save(_OWNER_ID, _master_profile())
    SqliteResumeVariantStore(_db_path()).save(
        _resume_variant(), user_id=_OWNER_ID
    )
    # A token minted for the wrong (non-owning) user must not unlock it.
    token = create_resume_download_token(
        user_id=_OTHER_ID,
        resume_variant_id="variant-1",
        secret_key=_JWT_SECRET,
        expires_in_days=90,
    )
    response = client.get(f"/export/resume/variant-1.pdf?token={token}")
    assert response.status_code == 404


def test_resume_pdf_download_rejects_an_unknown_variant(client: TestClient) -> None:
    _create_user(_OWNER_ID, "owner@example.com")
    token = create_resume_download_token(
        user_id=_OWNER_ID,
        resume_variant_id="no-such-variant",
        secret_key=_JWT_SECRET,
        expires_in_days=90,
    )
    response = client.get(f"/export/resume/no-such-variant.pdf?token={token}")
    assert response.status_code == 404


def test_export_never_leaks_another_users_rows(client: TestClient) -> None:
    owner_token = _create_user(_OWNER_ID, "owner@example.com")
    other_token = _create_user(_OTHER_ID, "other@example.com")
    SqliteApplicationSessionStore(_db_path()).save(
        _application_session(company="Owner Only Corp"), user_id=_OWNER_ID
    )

    owner_sheet = load_workbook(
        BytesIO(
            client.get(
                "/export/applications.xlsx",
                headers={"Authorization": f"Bearer {owner_token}"},
            ).content
        )
    ).active
    assert "Owner Only Corp" in [cell.value for cell in owner_sheet[2]]

    other_sheet = load_workbook(
        BytesIO(
            client.get(
                "/export/applications.xlsx",
                headers={"Authorization": f"Bearer {other_token}"},
            ).content
        )
    ).active
    assert other_sheet.max_row == 1  # the other user sees only the header
