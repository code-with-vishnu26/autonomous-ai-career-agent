"""Excel application-tracker export (Phase 13, ADR-0037; Phase 53, ADR-0071).

The original user requirement from the project's founding brief: one
command, one formatted, filterable openpyxl workbook of every recorded
application -- company, title, source, ATS score, truthfulness verdict,
status, dates, tier used, file paths, latest outcome. Reads the
append-only :class:`~career_agent.storage.sqlite.SqliteApplicationStore`;
never writes back to it.

:func:`export_submissions` (Phase 53) is a separate, sibling sheet-export
for :class:`~career_agent.domain.submission.SubmissionResult` -- kept
distinct from ``export_applications`` rather than merged into one table,
since the two pipelines (``apply``'s ``Application``/``SqliteApplicationStore``
vs. ``prepare``/``review``/``submit``'s ``ApplicationSession``/
``SubmissionResult``) are not the same rows and were never unified
(ADR-0069/0070/0071 each note this as a deliberate, separate storage
convention). A future combined view is named future work, not attempted
here.

:func:`export_application_sessions` (Phase 65, ADR-0083) adds a third,
sibling export for the dashboard's own ``ApplicationSession`` rows (what
``prepare`` writes and ``/api/applications`` returns), so a web user can
download the same "every application, in a spreadsheet" artifact the CLI
has always produced. The ``*_xlsx_bytes`` helpers serialize the identical
workbooks to memory instead of a path, for the ``/export/*.xlsx``
download endpoints -- the on-disk CLI output is unchanged, byte for byte.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from career_agent.domain.application_session import ApplicationSession
    from career_agent.domain.submission import SubmissionResult


def _build_workbook(
    title: str,
    columns: list[tuple[str, str]],
    rows: list[dict[str, object]],
    *,
    link_keys: frozenset[str] = frozenset(),
) -> Workbook:
    """Build a formatted, filterable single-sheet workbook.

    ``rows`` are dicts of *final* cell values keyed by each column's key --
    every caller does its own value transform (truthfulness verdict,
    yes/no booleans, joined lists) before calling this, so the shared
    formatting (bold frozen header, auto-filter, uniform width) lives in
    exactly one place. Extracted in Phase 65 (ADR-0083); the two original
    exports' output is unchanged.

    ``link_keys`` (Phase 69, ADR-0087) names columns whose non-empty string
    value is a public URL -- those cells become real clickable hyperlinks.
    A cell's value must be a plain ``http(s)`` URL for the link to attach;
    anything else is written as ordinary text, so a missing URL is just a
    blank, never a broken link.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title

    for column_index, (_key, label) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=label)
        cell.font = Font(bold=True)

    for row_index, row in enumerate(rows, start=2):
        for column_index, (key, _label) in enumerate(columns, start=1):
            value = row.get(key)
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if (
                key in link_keys
                and isinstance(value, str)
                and value.startswith(("http://", "https://"))
            ):
                cell.hyperlink = value
                cell.style = "Hyperlink"

    last_column = get_column_letter(len(columns))
    sheet.auto_filter.ref = f"A1:{last_column}{max(len(rows) + 1, 1)}"
    sheet.freeze_panes = "A2"
    for column_index in range(1, len(columns) + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 18

    return workbook


def _workbook_bytes(workbook: Workbook) -> bytes:
    """Serialize a workbook to ``.xlsx`` bytes for an HTTP download.

    ``openpyxl.Workbook.save`` accepts a file-like object, so this writes
    to an in-memory buffer -- the same bytes ``save(path)`` would write to
    disk, without touching the filesystem (the ``/export/*.xlsx`` routes
    stream these straight to the caller).
    """
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


_COLUMNS: list[tuple[str, str]] = [
    ("recorded_at", "Recorded"),
    ("company", "Company"),
    ("title", "Title / Summary"),
    ("source", "Source"),
    ("status", "Status"),
    ("truthfulness_approved", "Truthfulness"),
    ("ats_total", "ATS Score"),
    ("tier_used", "Tier"),
    ("profile_version", "Profile Version"),
    ("prompt_version", "Prompt Version"),
    ("artifact_paths", "Files"),
    ("latest_outcome", "Latest Outcome"),
]


def _applications_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    """Render ``SqliteApplicationStore.all_rows`` for the sheet.

    Only the boolean-ish ``truthfulness_approved`` is transformed to a
    human-readable verdict; every other value is written as-is (ATS score
    stays a number, etc.), the exact behavior since Phase 13.
    """
    prepared: list[dict[str, object]] = []
    for row in rows:
        rendered = dict(row)
        rendered["truthfulness_approved"] = (
            "approved" if row.get("truthfulness_approved") else "blocked"
        )
        prepared.append(rendered)
    return prepared


def export_applications(rows: list[dict[str, object]], path: Path) -> Path:
    """Write ``rows`` (from ``SqliteApplicationStore.all_rows``) to ``path``.

    Returns the written path. Boolean-ish integers are rendered as
    approved/blocked so a human reads the sheet without decoding; the
    header row is bold and frozen, and an auto-filter spans the table --
    "formatted, filterable" per the founding requirement.
    """
    workbook = _build_workbook("Applications", _COLUMNS, _applications_rows(rows))
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def applications_xlsx_bytes(rows: list[dict[str, object]]) -> bytes:
    """The same workbook as :func:`export_applications`, as bytes."""
    return _workbook_bytes(
        _build_workbook("Applications", _COLUMNS, _applications_rows(rows))
    )


_SUBMISSION_COLUMNS: list[tuple[str, str]] = [
    ("submitted_at", "Submitted"),
    ("company", "Company"),
    ("job_title", "Role"),
    ("provider", "Provider"),
    ("status", "Status"),
    ("submitted", "Submitted?"),
    ("confirmation_id", "Confirmation ID"),
    ("duration_seconds", "Duration (s)"),
    ("refusal_reason", "Refusal Reason"),
    ("warnings", "Warnings"),
]


def _submissions_rows(results: list[SubmissionResult]) -> list[dict[str, object]]:
    """Render ``SubmissionResult`` objects for the sheet.

    Every value is stringified (``str(value or "")``) and ``submitted`` is
    rendered yes/no, the exact behavior since Phase 53 -- so a blank cell
    reads as blank rather than the string ``"None"``.
    """
    prepared: list[dict[str, object]] = []
    for result in results:
        row = result.model_dump()
        row["warnings"] = "; ".join(result.warnings)
        rendered: dict[str, object] = {}
        for key, _label in _SUBMISSION_COLUMNS:
            value = row.get(key)
            if key == "submitted":
                value = "yes" if value else "no"
            rendered[key] = str(value or "")
        prepared.append(rendered)
    return prepared


def export_submissions(results: list[SubmissionResult], path: Path) -> Path:
    """Write ``results`` to ``path``.

    From ``SqliteSubmissionResultStore.all_results`` -- same formatted/
    filterable shape as ``export_applications`` (Phase 53, ADR-0071), a
    separate sheet/file since the two pipelines track different rows.
    """
    workbook = _build_workbook(
        "Submissions", _SUBMISSION_COLUMNS, _submissions_rows(results)
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def submissions_xlsx_bytes(results: list[SubmissionResult]) -> bytes:
    """The same workbook as :func:`export_submissions`, as bytes."""
    return _workbook_bytes(
        _build_workbook("Submissions", _SUBMISSION_COLUMNS, _submissions_rows(results))
    )


_APPLICATION_SESSION_COLUMNS: list[tuple[str, str]] = [
    ("created_at", "Prepared"),
    ("company", "Company"),
    ("job_title", "Role"),
    ("provider", "Provider / ATS"),
    ("status", "Status"),
    ("resume_variant_id", "Résumé Variant"),
    ("cover_letter", "Cover Letter"),
    ("fields_filled", "Fields Filled"),
    ("fields_missing", "Fields Missing"),
    ("uploaded_files", "Files Uploaded"),
    ("url", "Job URL"),
    ("warnings", "Warnings"),
]


def _application_sessions_rows(
    sessions: list[ApplicationSession],
) -> list[dict[str, object]]:
    """Render ``ApplicationSession`` objects for the dashboard export.

    The list-valued fields (``filled_fields``/``missing_fields``/
    ``uploaded_files``/``warnings``) become human-readable counts or
    joined text -- a spreadsheet cell can't usefully hold a Python list,
    and a reviewer scanning "how far did preparation get?" wants the
    count and the missing-field names, not raw selectors.
    """
    prepared: list[dict[str, object]] = []
    for session in sessions:
        prepared.append(
            {
                "created_at": session.created_at.isoformat(),
                "company": session.company,
                "job_title": session.job_title,
                "provider": session.provider,
                "status": str(session.status),
                "resume_variant_id": session.resume_variant_id or "",
                "cover_letter": "yes" if session.cover_letter_body else "no",
                "fields_filled": len(session.filled_fields),
                "fields_missing": "; ".join(session.missing_fields),
                "uploaded_files": "; ".join(session.uploaded_files),
                "url": session.url,
                "warnings": "; ".join(session.warnings),
            }
        )
    return prepared


def export_application_sessions(
    sessions: list[ApplicationSession], path: Path
) -> Path:
    """Write prepared ``ApplicationSession`` rows to ``path`` (Phase 65).

    The dashboard analogue of :func:`export_applications` -- the ``apply``
    pipeline's ``Application`` rows and ``prepare``'s ``ApplicationSession``
    rows are genuinely different records (see the module docstring), so
    this is a third sibling export, not a merge.
    """
    workbook = _build_workbook(
        "Applications",
        _APPLICATION_SESSION_COLUMNS,
        _application_sessions_rows(sessions),
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def application_sessions_xlsx_bytes(sessions: list[ApplicationSession]) -> bytes:
    """The same workbook as :func:`export_application_sessions`, as bytes."""
    return _workbook_bytes(
        _build_workbook(
            "Applications",
            _APPLICATION_SESSION_COLUMNS,
            _application_sessions_rows(sessions),
        )
    )


#: Enriched application export (Phase 69, ADR-0087). The router assembles a
#: dict per application joining the ApplicationSession with its Opportunity
#: (accurate posting details + public company links) and optional
#: web-search company research; every value here is already a finished
#: cell (string/number), so this module needs no domain imports for it.
_ENRICHED_APPLICATION_COLUMNS: list[tuple[str, str]] = [
    ("prepared", "Prepared"),
    ("company", "Company"),
    ("role", "Role"),
    ("location", "Location"),
    ("remote", "Remote"),
    ("source", "Source"),
    ("posted", "Posted"),
    ("status", "Status"),
    ("job_url", "Job URL"),
    ("careers_url", "Careers Page"),
    ("linkedin_url", "Company LinkedIn"),
    ("company_research", "Company Research"),
    ("research_sources", "Research Sources"),
    ("resume_pdf_url", "Résumé (PDF)"),
    ("cover_letter", "Cover Letter"),
]

#: Columns whose values are single URLs -> clickable hyperlinks. The last
#: three are public; ``resume_pdf_url`` (Phase 71, ADR-0089) is instead a
#: signed, capability-bearing link scoped to exactly the caller's own
#: tailored résumé -- still a single URL, so it renders as a hyperlink the
#: same way.
_ENRICHED_LINK_KEYS = frozenset(
    {"job_url", "careers_url", "linkedin_url", "resume_pdf_url"}
)


def enriched_applications_xlsx_bytes(rows: list[dict[str, object]]) -> bytes:
    """An enriched applications workbook (Phase 69, ADR-0087).

    ``rows`` are pre-assembled by the export router (session + opportunity +
    optional research), so this stays a thin formatting call -- the same
    ``_build_workbook`` as every other export, with the URL columns turned
    into real hyperlinks.
    """
    return _workbook_bytes(
        _build_workbook(
            "Applications",
            _ENRICHED_APPLICATION_COLUMNS,
            rows,
            link_keys=_ENRICHED_LINK_KEYS,
        )
    )
